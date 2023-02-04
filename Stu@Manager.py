from tkinter import *
from datetime import date
from tkinter import filedialog
from tkinter import messagebox
from PIL import Image , ImageTk


import os
from tkinter.ttk import Combobox
import openpyxl , xlrd 
from openpyxl import Workbook
import pathlib


background="#06283D"
framebg="#EDEDED"
framefg="#06283D"



root = Tk()
root.title("Studen Registration System")
root.geometry("1250x700+180+60")
root.config(bg=background)
icon = PhotoImage(file='./photo/logo.png')
root.iconphoto(True,icon)

file = pathlib.Path('Student_data.xlsx')

if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']="Registration No."
    sheet['B1']="Name"
    sheet['C1']="Class"
    sheet['D1']="Gender"
    sheet['E1']="DOB"
    sheet['F1']="Date Of Registration"
    sheet['G1']="Nationality"
    sheet['H1']="Skills"
    sheet['I1']="Grade"
    sheet['N1']="Email : "
    sheet['J1']="Father Name"
    sheet['K1']="Mother Name"
    sheet['L1']="Father's Occupation"
    sheet['M1']="Mother's Occcupation"
    
    
    
    file.save('Student_data.xlsx')
    
#exit window ##############################################
def Exit():
    root.destroy()
##################################################SHow image   
def showImage():
    global filename
    global img
    filename=filedialog.askopenfilename(initialdir=os.getcwd(),
                                        title="Select image file",filetype=(("JPG FILE","*.jpg"),
                                                                            ("PNG FILE","*.png"),
                                                                            ("ALL files","*.txt")))
    img=(Image.open(filename))
    resized_image= img.resize((190,190))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2 
    
############################### REGISTRATION No. #######################
# lets desing automatic registration n

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row=sheet.max_row
    
    max_row_value=sheet.cell(row=row,column=1).value
    
    try:
        Registration.set(max_row_value+1)
    except:
        Registration.set("1")

###################    Clear  ######################
def Clear():
    Name.set('')
    DOB.set('')
    Nationality.set('')
    Skills.set('')
    F_Name.set('')
    M_Name.set('')
    F_Occupation.set('')
    M_Occupation.set('')
    Grade.set('')
    Email.set('')
    Class.set('Select Class')
    
    registration_no()
    
    saveButton.config(state= 'normal')
    img1=PhotoImage(file='./photo/uploadlogocopie.png')
    lbl.config(image=img1)
    lbl.image=img1
    
    img=""
    
###################### 
def save() :
        R1=Registration.get()
        N1=Name.get()
        C1=Class.get()
        try:
            G1 = gender
        except:
            messagebox.showerror("error","Select Gender !")
        D2=DOB.get()
        D1=Date.get()
        S1=Skills.get()
        Re1 = Nationality.get()
        Gr=Grade.get()
        E1=Email.get()
        
        fathername=F_Name.get()
        mothername=M_Name.get()
        F1=F_Occupation.get()
        M1=M_Occupation.get()
        
        
        if N1=="" or C1=="Select Class" or D2=="" or Re1=="" or S1=="" or Re1=="" or E1=="":
            messagebox.showerror("error","Few Data is messing!")
            
        else:
            file = openpyxl.load_workbook('Student_data.xlsx')     
            sheet = file.active
            sheet.cell(column=1,row=sheet.max_row+1,value=R1)  
            sheet.cell(column=2,row=sheet.max_row,value=N1)
            sheet.cell(column=3,row=sheet.max_row,value=C1)
            sheet.cell(column=4,row=sheet.max_row,value=G1)
            sheet.cell(column=5,row=sheet.max_row,value=D2)
            sheet.cell(column=6,row=sheet.max_row,value=D1)
            sheet.cell(column=7,row=sheet.max_row,value=Re1)
            sheet.cell(column=8,row=sheet.max_row,value=S1)
            sheet.cell(column=9,row=sheet.max_row,value=Gr)
            
            sheet.cell(column=10,row=sheet.max_row,value=fathername)
            sheet.cell(column=11,row=sheet.max_row,value=mothername)
            sheet.cell(column=12,row=sheet.max_row,value=F1)
            sheet.cell(column=13,row=sheet.max_row,value=M1)
            sheet.cell(column=14,row=sheet.max_row,value=E1)
            file.save('Student_data.xlsx')
        
            try:
                img.save("Student pic/"+str(R1)+".jpg")
            except:
                messagebox.showerror("info","Profile picture is not available !!!")
                
            messagebox.showinfo("info"," Data entered with succes !")
            
            Clear()
            registration_no()
        
###################### Search ################        
def search():
    
    global x1
    text = Search.get()
    
    
    Clear()
    saveButton.config(state='disable') #no one can click on it
    file=openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active

    for row in sheet.rows:
        if row[0].value == int(text): 
            name=row[0]
       # print(str(name))   
            reg_no_position = str(name)[14:-1]   # ex column 1
            reg_number=str(name)[15:-1]          # A2
    try:
        print(str(name))
    except:
        messagebox.showerror("Invalid","Invalid Registration Number !!!")
        
    #reg_no_position showing like A1, A2
    x1 = sheet.cell(row=int(reg_number),column=1).value
    x2 = sheet.cell(row=int(reg_number),column=2).value
    x3 = sheet.cell(row=int(reg_number),column=3).value
    x4 = sheet.cell(row=int(reg_number),column=4).value
    x5 = sheet.cell(row=int(reg_number),column=5).value
    x6 = sheet.cell(row=int(reg_number),column=6).value
    x7 = sheet.cell(row=int(reg_number),column=7).value
    x8 = sheet.cell(row=int(reg_number),column=8).value
    x9=  sheet.cell(row=int(reg_number),column=9).value
    x10= sheet.cell(row=int(reg_number),column=10).value
    x11= sheet.cell(row=int(reg_number),column=11).value
    x12= sheet.cell(row=int(reg_number),column=12).value
    x13= sheet.cell(row=int(reg_number),column=13).value
    x14= sheet.cell(row=int(reg_number),column=14).value
    
    Registration.set(x1)
    Name.set(x2)
    Class.set(x3) 
    if x4=='Male':
        R1.select()
    else:
        R2.select()
    
    DOB.set(x5)
    Date.set(x6)
    Skills.set(x8)
    Nationality.set(x7)
    Grade.set(x9)
    F_Name.set(x10)
    M_Name.set(x11)
    F_Occupation.set(x12)
    M_Occupation.set(x13)
    Email.set(x14)
    
    img=(Image.open("./Student pic/"+str(x1)+".jpg"))
    
    resized_image = img.resize((180,180))
    photo2 = ImageTk.PhotoImage(resized_image)
    lbl.config(image=photo2)
    lbl.image=photo2
    
################# UPDATE ###############################
def Update():
    R1=Registration.get()
    N1=Name.get()
    C1=Class.get()
    selection()
    G1=gender.get()  
    D2=DOB.get()
    D1=Date.get()
    S1=Skills.get()
    Re1 = Nationality.get()
    Gr=Grade.get()
    E1=Email.get()
        
    fathername=F_Name.get()
    mothername=M_Name.get()
    F1=F_Occupation.get()
    M1=M_Occupation.get()
    
    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet=file.active
    for row in sheet.rows:
        if row[0].value== R1:
            name=row[0]
            print(str(name))
            reg_no_position=str(name)[14:-1]
            reg_number=str(name)[15:-1]
    #sheet.cell(column=1,row=int(reg_number),value=R1)
    sheet.cell(column=2,row=int(reg_number),value=N1)
    sheet.cell(column=3,row=int(reg_number),value=C1)
    sheet.cell(column=4,row=int(reg_number),value=G1)
    sheet.cell(column=5,row=int(reg_number),value=D2)
    sheet.cell(column=6,row=int(reg_number),value=D1)
    sheet.cell(column=7,row=int(reg_number),value=Re1)
    sheet.cell(column=8,row=int(reg_number),value=S1)
    sheet.cell(column=9,row=int(reg_number),value=Gr)
    sheet.cell(column=10,row=int(reg_number),value=fathername)
    sheet.cell(column=11,row=int(reg_number),value=mothername)
    sheet.cell(column=12,r2ow=int(reg_number),value=F1)
    sheet.cell(column=13,ro3w=int(reg_number),value=M1)
    sheet.cell(column=14,row=int(reg_number),value=E1)
    file.save("Student_data.xlsx")
    try:
        img.save("./Student pic/"+str(x1)+".jpg")
    except:
        pass
    messagebox.showinfo("Update","Update Successfully")
    Clear()
    
#gender

def selection():
    global gender
    value=radio.get() 
    if value==1 :
        gender="Male"
       
    else :
        gender ="Female"
        
        
        
        

#top frames
Label(root,text="Created by GLEI JIHED .. EMAIL : glei.jihed@gmail.com",width=10,height=3,bg="#f0687C",anchor='e').pack(side=TOP,fill=X)
Label(root,text="STUDENT REGISTRATION ",width=10,height=2,fg='#fff',font='arial 20 bold',bg="#c36464").pack(side=TOP,fill=X)

#search box to update

Search=StringVar()
Entry(root,textvariable=Search,width=15,bd=2,font="arial 20").place(x=830,y=70)
imageicon3 = PhotoImage(file=".\photo\searchcopie1.png")
Srch=Button(root,text="Search",compound=LEFT,image=imageicon3,width=137,height=30,bg='#68ddfa',font='arial 13 bold',command=search)
Srch.place(x=1067,y=70)

imageicon4=PhotoImage(file=".\photo\layercopie.png")
update_button=Button(root,image=imageicon4,bg='#c36464',command=Update)
update_button.place(x=110,y=62)

#registration and date
Label(root,text="Registration No:",font="arial 13",fg=framebg,bg=background).place(x=40,y=130)
Label(root,text="DATE:",font="arial 13",fg=framebg,bg=background).place(x=500,y=130)

Date=StringVar()
Registration = IntVar()

reg_enrty = Entry(root,textvariable=Registration,width=15,font="arial 10")
reg_enrty.place(x=170,y=133)

registration_no() 

#registration_no()
today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root,textvariable= Date ,width=15,font="arial 10")
date_entry.place(x=560,y=132)
Date.set(d1)

#Student details 
obj = LabelFrame(root,text="Student's Detail's",font=20,bd=2,bg=framebg,fg=framefg,width=900,height=250,relief=GROOVE)
obj.place(x=30,y=200)

Label(obj,text="Full Name: ",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=40)
Label(obj,text="Date Of Birth : ",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=80)
Label(obj,text="Gender : ",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=120)
Label(obj,text="Grade :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=162)

Label(obj,text="Class : ",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=40)
Label(obj,text="Nationality : ",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=80)
Label(obj,text="Skills : ",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=120)
Label(obj,text="Email :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=162)

Name = StringVar()
name_entry = Entry(obj,textvariable=Name,width=20,font="arial 10")
name_entry.place(x=160,y=43)

DOB = StringVar()
dob_entry = Entry(obj,textvariable = DOB ,width=20,font="arial 10")
dob_entry.place(x=160,y=83)



radio=IntVar()
R1 = Radiobutton(obj,text="Male",variable=radio,value=1,bg=framebg,fg=framefg,command=selection)
R1.place(x=152,y=122)

R2 = Radiobutton(obj,text="Female",variable=radio,value=2,bg=framebg,fg=framefg,command=selection)
R2.place(x=202,y=122)

Nationality=StringVar()
natio_entry = Entry(obj,textvariable=Nationality,width=20,font="arial 10")
natio_entry.place(x=630,y=83)

Skills=StringVar()
skills_entry = Entry(obj,textvariable=Skills,width=20,font="arial 10")
skills_entry.place(x=630,y=123)

Email=StringVar()
email_entry = Entry(obj,textvariable=Email,width=30,font="arial 10")
email_entry.place(x=630,y=163)


Grade=IntVar()
grade_entry=Entry(obj,textvariable=Grade,width=20,font="arial10")
grade_entry.place(x=150,y=163)

Class =  Combobox(obj,values=['B1','B2','B3','M1','M2','MS','ING1','ING2','ING3'],font="Roboto 10",width=17,state="r")
Class.place(x=630,y=43)
Class.set("Select Class")


#Parents details:

obj2 = LabelFrame(root,text="Parent's Detail's",font=20,bd=2,bg=framebg,fg=framefg,width=900,height=220,relief=GROOVE)
obj2.place(x=33,y=470)

Label(obj2,text="Father's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=40)
Label(obj2,text="Occupation :",font="arial 13",bg=framebg,fg=framefg).place(x=30,y=80)

Label(obj2,text="Mother's Name:",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=40)
Label(obj2,text="Occupation :",font="arial 13",bg=framebg,fg=framefg).place(x=500,y=80)


F_Name = StringVar()
f_entry = Entry(obj2,textvariable=F_Name,width=20,font="arial 10")
f_entry.place(x=160,y=43)

F_Occupation = StringVar()
fo_entry = Entry(obj2,textvariable=F_Occupation,width=20,font="arial 10")
fo_entry.place(x=160,y=83)

M_Name = StringVar()
m_entry = Entry(obj2,textvariable=M_Name,width=20,font="arial 10")
m_entry.place(x=630,y=43)

M_Occupation = StringVar()
mo_entry = Entry(obj2,textvariable=M_Occupation,width=20,font="arial 10")
mo_entry.place(x=630,y=83)

#image

f = Frame(root,bd=3,bg="black",width=180,height=180,relief=GROOVE)
f.place(x=1000,y=150)

img=PhotoImage(file="./photo/uploadlogocopie.png")
lbl=Label(f,bg="black",image=img)
lbl.place(x=0,y=0)

#button
Button(root,text="UPLOAD",width=19,height=2,font="arial 12 bold",bg="lightblue",command=showImage).place(x=1000,y=370)

saveButton=Button(root,text="SAVE",width=19,height=2,font="arial 12 bold",bg="lightgreen",command=save)
saveButton.place(x=1000,y=450)
Button(root,text="RESET",width=19,height=2,font="arial 12 bold",bg="lightpink",command=Clear).place(x=1000,y=530)
Button(root,text="EXIT",width=19,height=2,font="arial 12 bold",bg="lightgrey",command=Exit).place(x=1000,y=610)












































root.mainloop()